#!/usr/bin/which python3
# 3.apply-formulas.py
# loads the Excel pivot_table.xlsx 'Report' workbook created in 1.make-pivot-table.py
# select the 'Report' sheet to generate the graph
# select 3 columns by Gender, Product, Total Sales
# create a pivot table using Gender for y-axis showing Product and Total Sales
# output to excel
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import sys
PROG = os.path.basename(sys.argv[0])
IN_FILE = os.path.expanduser('barchart.xlsx')
OUT_FILE = os.path.expanduser('report.xlsx')

if not os.path.exists(IN_FILE):
    print("{} -- file '{}' not found".format(PROG, IN_FILE))
    exit(1)

# if the barchart.xlsx file has been modified by IntelliJ's ExcelReader,
# this will throw a 'KeyError' exception
try:
    wb = load_workbook(IN_FILE)
except KeyError:
    print("{} -- error opening '{}'... regenerate the file".format(PROG, IN_FILE))
    exit(1)

sheet = wb['Report']
print("{}--> '{}({})'".format(PROG, IN_FILE, sheet), end='', flush=True)

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# this is what this code does if you did it manually
# Write an Excel formula with Python
# sheet['B8'] = '=SUM(B6:B7)'
# sheet['B8'].style = 'Currency'

# Write multiple formulas with a for loop
for i in range(min_column+1, max_column+1):  # (B, G+1) skipping category
    letter = get_column_letter(i)            # conv column # to letter
    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency' # formats as "####.## $"

wb.save(OUT_FILE)
print(" --> '{}'".format(OUT_FILE))
