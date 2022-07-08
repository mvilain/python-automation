#!/usr/bin/which python3
"""
6.py-to-exe.py
loads the Excel the sales marketing data and creates a pivot_table.xlsx 'Report' workbook
select the 'Report' sheet
create bar chart
create total summary of categories
annotate cells A1 and A2 with titles
output to excel from openpyxl import load_workbook
"""

import argparse
import os
import sys
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def parse_arguments(default_input_file='supermarket_sales.xlsx',
                    default_output_file='report.xlsx',
                    default_month='MONTHLY'
                    ):
    """parse the argument list, build help and usage messages

    Args:
        default_input_file (str): filename containing a default input Excel spreadsheet
        default_output_file (str): filename containing a default output Excel spreadsheet for
        the consolidation report
        default_month (str): containing the current month as a string (not validated)

    Returns:
        namespace (ns): namespace with the arguments passed and their values

    """
    parser = argparse.ArgumentParser(
        description='generate Consolidation report from Excel spreadsheet of marketing data')
    # if omitted, the default value is returned, so arg.input is always defined
    parser.add_argument('-i', '--input',
                        action='store',
                        default=default_input_file,
                        help=f'Input file of Excel Marketing data [default: {default_input_file}]',
                        required=False,
                        # nargs="?",   # command line arg w/o flag
                        )
    # if omitted, the default value is returned, so arg.output is always defined
    parser.add_argument('-o', '--output',
                        action="store",
                        default=default_output_file,
                        help=f'Consolidation Report output file [default: {default_output_file}]',
                        required=False
                        )
    # if omitted, the current month is used, so arg.month will always be defined
    parser.add_argument('-m', '--month',
                        action="store",
                        default=default_month,
                        help=f'Month to tag in output report [default: {default_month}]',
                        required=False
                        )
    parser.add_argument('-v', '--verbose',
                        action="store_true",
                        help='show progress as report is being built',
                        required=False
                        )
    args = parser.parse_args()

    # args.input = os.path.expanduser(args.input)
    # args.output = os.path.expanduser(args.output)

    return args  # namespace containing the argument passed on the command line


def main():
    """main program to input marketing data and output a consolidation report as Excel spreadsheet

    Args:
        inputs Excel spreadsheet (must be xlsx file)
        assumes Excel table format with 1st line containing
    - Invoice
    - Branch
    - City
    - Customer Type
    - Gender
    - Product Line
    - Unit Price
    - Quantity
    - Tax 5%
    - Total
    - Date (mm/dd/yy)
    - Time (HH:MM)
    - Payment
    - cogs
    - gross Margin
    - gross Income
    - Rating
        outputs Excel spreadsheet consolidation report

    Returns: 0

"""
    prog = os.path.basename(sys.argv[0])
    this_month = date.today().strftime('%B')  # full month
    args = parse_arguments(
                            default_input_file='supermarket_sales.xlsx',
                            default_output_file=f'report_{this_month}.xlsx',
                            default_month=this_month
                           )

    input_expanded = os.path.expanduser(args.input)
    # output_file = os.path.expanduser(os.path.exists(args.output))
    if not os.path.exists(input_expanded):
        print(f'{prog} -- file {input_expanded} not found')
        sys.exit(1)

    df = pd.read_excel(input_expanded)  # open excel file as panda dataframe

    if args.verbose:
        print(f'{prog}--> read_excel {input_expanded}...', end='', flush=True)

    # Select columns: 'Gender', 'Product line', 'Total'
    df = df[['Gender', 'Product line', 'Total']]
    if args.verbose:
        # print(df)
        print('selecting columns...', end='', flush=True)

    pivot_table = df.pivot_table(index='Gender', columns='Product line',
                                 values='Total', aggfunc='sum').round(2)

    # Export pivot table to Excel file w/ sheet starting at row 4
    # Preparing script before we convert it to executable
    # application_path = os.path.dirname(sys.executable)
    # output_file = os.path.join(os.path.dirname(sys.executable), args.output)
    pivot_table.to_excel(args.output, 'Report', startrow=4)
    if args.verbose:
        # print(pivot_table)
        print(f'excel.pivot{args.output}...')

    # if the pivot_table.xlsx file has been modified by IntelliJ's ExcelReader,
    # this will throw a 'KeyError' exception, so trap it and handle that
    # NOTE: opening the file with Apple's Numbers or Excel does not cause this error
    # since this is an intermediate step and can't be read, remove the exception trapping
    # try:
    wb = load_workbook(args.output)
    # except KeyError:
    #     print(f"{PROG} -- error opening '{output_file}'... regenerate the file")
    #     sys.exit(1)
    #
    # TODO: how to reference the sheet from report_sheet as a dict w/o hard coding reference?
    try:
        sheet = wb['Report']  # not defined if Sheet not found...throws KeyError
        if args.verbose:
            print(f"{prog}--> '{args.output}(Report)'...", end='', flush=True)
    except KeyError:
        print(f"{prog} -- error opening '{args.output}' -- workbook or sheet not found")
        sys.exit(1)

    if args.verbose:
        print('...loaded...', end='', flush=True)

    # Active rows and columns
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # Instantiate a barchart
    barchart = BarChart()

    # Locate data and categories
    data = Reference(sheet,
                     min_col=min_column+1, max_col=max_column,  # data starts after header column
                     min_row=min_row,      max_row=max_row)
    categories = Reference(sheet,
                           min_col=min_column, max_col=min_column,  # omit header column
                           min_row=min_row+1,  max_row=max_row)

    # Adding data and categories
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)

    # Make chart in cell B12
    sheet.add_chart(barchart, "B12")
    barchart.title = 'Sales by Product line'
    # chart style for Excel 16.6 on MacOS 12.4
    # plain    1=BW  2=multi-color  3=blue  4=red  5=green  6=purple  7=cyan  8=orange
    # outlined 9=BW 10=multi-color 11=blue 12=red 13=green 14=purple 15=cyan 16=orange
    # no error occurs of style > 16
    barchart.style = 5  # choose the chart style (plain green)
    if args.verbose:
        print('barchart...', end='', flush=True)
    # Write multiple formulas with a for loop
    # note that openpyxl has a bug incorrectly setting style='Currency' (##### $ instead of $#####)
    for i in range(min_column+1, max_column+1):  # (B, G+1)
        letter = get_column_letter(i)
        sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
        sheet[f'{letter}{max_row + 1}'].style = 'Currency'
    if args.verbose:
        print('summary...', end='', flush=True)

    # Add Title and format
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = args.month
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)
    if args.verbose:
        print('titles...', end='', flush=True)

    wb.save(args.output)
    if args.verbose:
        print(f" '{args.output}'")

    return 0


if __name__ == '__main__':
    sys.exit(main())
