#!/usr/bin/which python3
# 2.add-charts.py
# loads the Excel pivot_table.xlsx 'Report' workbook created in 1.make-pivot-table.py
# select the 'Report' sheet to generate the graph
# select 3 columns by Gender, Product, Total Sales
# create a pivot table using Gender for y-axis showing Product and Total Sales
# output to excel

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import os
import sys
PROG = os.path.basename(sys.argv[0])
IN_FILE = os.path.expanduser('pivot_table.xlsx')
OUT_FILE = os.path.expanduser('barchart.xlsx')

# Read workbook and select sheet
if not os.path.exists(IN_FILE):
    print("{} -- file '{}' not found".format(PROG, IN_FILE))
    exit(1)

# if the pivot_table.xlsx file has been modified by IntelliJ's ExcelReader,
# this will throw a 'KeyError' exception, so trap it and handle that
# NOTE: opening the file with Apple's Numbers or Excel does not cause this error
try:
    wb = load_workbook(IN_FILE)
except KeyError:
    print("{} -- error opening '{}'... regenerate the file".format(PROG, IN_FILE))
    exit(1)

try:
    sheet = wb['Report']  # not defined if Sheet not found...throws KeyError
    print("{}--> '{}({})'".format(PROG, IN_FILE, sheet), end='', flush=True)
except KeyError:
    print("{} -- error opening '{}' -- workbook or sheet not found".format(PROG, IN_FILE))
    exit(1)

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row
# print('  [active cells=({},{}):({},{})] '.format(min_row,min_column,max_row,max_column))

# Instantiate a barchart
barchart = BarChart()

# Locate data and categories
# data reference omits the column with the category headers
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row
                 )
# categories being displayed are in Column A (e.g.1)
# don't include the headers in min_row
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row
                       )

# Make chart
# Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

barchart.title = 'Sales by Product line'
# chart style
# plain   1=BW  2=multi-color 3=blue  4=red  5=green  6=purple  7=cyan  8=orange
# outline 9=BW 10=multi-color 11=blue 12=red 13=green 14=purple 15=cyan 16=orange
# no error occurs of style > 16
barchart.style = 5  # choose the chart style
sheet.add_chart(barchart, "B12")

# Save workbook
wb.save(OUT_FILE)
print(" --> '{}'".format(OUT_FILE))
