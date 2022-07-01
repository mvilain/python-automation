#!/usr/bin/which python3
# 5.pivot-to-report.py
# loads the Excel pivot_table.xlsx 'Report' workbook created in 1.make-pivot-table.py
# select the 'Report' sheet
# create bar chart
# create total summary of categories
# annotate cells A1 and A2 with titles
# output to excel
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import sys
PROG = os.path.basename(sys.argv[0])
MONTH = 'february'
IN_FILE = os.path.expanduser('pivot_table.xlsx')
OUT_FILE = os.path.expanduser(f'report_{MONTH}.xlsx')

# Read workbook and select sheet
if not os.path.exists(IN_FILE):
    print("{} -- file '{}' not found".format(PROG, IN_FILE))
    exit(1)

# Putting together #2, #3, and #4 (input: pivot_table.xlsx + month , output: Report with barchart, formulas and format)

# if the pivot_table.xlsx file has been modified by IntelliJ's ExcelReader,
# this will throw a 'KeyError' exception, so trap it and handle that
# NOTE: opening the file with Apple's Numbers or Excel does not cause this error
try:
    wb = load_workbook(IN_FILE)
except KeyError:
    print("{} -- error opening '{}'... regenerate the file".format(PROG, IN_FILE))
    exit(1)

try:
    sheet = wb['Report']  # not defined if Sheet not found...throws KeyError
    print("{}--> '{}({})'".format(PROG, IN_FILE, sheet), end='', flush=True)
except KeyError:
    print("{} -- error opening '{}' -- workbook or sheet not found".format(PROG, IN_FILE))
    exit(1)

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Instantiate a barchart
barchart = BarChart()

# Locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row
                 )  # including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row
                       )  # not including headers

# Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Make chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
# chart style for Excel 16.6 on MacOS
# plain   1=BW  2=multi-color 3=blue  4=red  5=green  6=purple  7=cyan  8=orange
# outline 9=BW 10=multi-color 11=blue 12=red 13=green 14=purple 15=cyan 16=orange
# no error occurs of style > 16
barchart.style = 5  # choose the chart style

# Write multiple formulas with a for loop
for i in range(min_column+1, max_column+1):  # (B, G+1)
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    # Python's 'Currency' style appears as "##### $" which shows as Custom in Excel
    # this is a bug in the openpyxl library
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'

# Add format
sheet['A1'] = 'Sales Report'
sheet['A2'] = MONTH
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

wb.save(OUT_FILE)
print(" --> '{}'".format(OUT_FILE))
